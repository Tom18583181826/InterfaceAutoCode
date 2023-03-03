import requests
from openpyxl import load_workbook
# openpyxl是Python中用于处理Excel表格的第三方类库
from InterfaceAutoTest.common.read_ini import ReadIni
from InterfaceAutoTest.common.read_json import read_json
from InterfaceAutoTest.data.excel_column import ExcelColumn


class ReadExcel:
    def __init__(self):
        self.read_ini = ReadIni()
        # 使用ReadIni()生成read_ini对象
        self.excel_path = self.read_ini.get_excel_path()
        # 通过read_ini对象调用get_excel_path()方法获取用例文件路径
        self.params_file_path = self.read_ini.get_params_file_path()
        # 通过read_ini对象调用get_params_file_path()方法获取参数文件路径
        self.expect_file_path = self.read_ini.get_expect_file_path()
        # 通过read_ini对象调用get_expect_file_path()方法获取预期结果文件路径
        self.wb = load_workbook(self.excel_path)
        # 传入excel表格路径获取excel表格
        self.wb_sheet = self.wb[self.read_ini.get_sheet_name()]
        # 通过read_ini对象调用get_sheet_name()方法获取sheet页名称，通过名称指定所需要的sheet页

    # 获取指定单元格的值,传入的参数值为excel表的列和行
    def get_cell_value(self, column, row):
        return self.wb_sheet[column + str(row)].value
        # 将行数字类型转换成字符串类型，使用value获取单元格中的值

    # 获取用例编号,传入的参数值为excel表的行，列指定写死
    def get_case_id(self, row):
        return self.get_cell_value(ExcelColumn.CASE_ID, row)

    # 获取用例标题
    def get_case_name(self, row):
        return self.get_cell_value(ExcelColumn.CASE_NAME, row)

    # 获取请求方法
    def get_case_method(self, row):
        return self.get_cell_value(ExcelColumn.CASE_METHOD, row)

    # 获取请求URL
    def get_case_url(self, row):
        return self.get_cell_value(ExcelColumn.CASE_URL, row)

    # 获取请求参数的key
    def get_case_params_key(self, row):
        return self.get_cell_value(ExcelColumn.CASE_PARAMS_KEY, row)

    # 获取预期结果的key
    def get_case_expect_key(self, row):
        return self.get_cell_value(ExcelColumn.CASE_EXPECT_KEY, row)

    # 获取请求参数的value
    def get_case_params_value(self, row):
        if self.get_case_params_key(row):
            # 当excel表中的请求参数key不为空时，调用read_json方法读取请求参数文件中的value值
            return read_json(self.params_file_path)[self.get_case_params_key(row)]
            # 传入参数为文件路径，[]中指定具体读取的参数key值

    # 获取预期结果的value
    def get_case_expect_value(self, row):
        if self.get_case_expect_key(row):
            return read_json(self.expect_file_path)[self.get_case_expect_key(row)]

    # 获取excel表的行数
    def get_row_count(self):
        return self.wb_sheet.max_row

    # 获取excel表的列数
    def get_column_count(self):
        return self.wb_sheet.max_column

    # 获取前置用例的id
    def get_case_precondition_id(self, row):
        return self.get_cell_value(ExcelColumn.CASE_PRECONDITION_ID, row)
        # 调用获取单元格的值的方法，传入列和行，获取前置用例的id

    # 获取依赖字段
    def get_case_depend_key(self, row):
        return self.get_cell_value(ExcelColumn.CASE_DEPEND_KEY, row)

    # 获取正则表达式
    def get_case_re(self, row):
        return self.get_cell_value(ExcelColumn.CASE_RE, row)

    # 获取是否执行的标志
    def get_if_execute(self, row):
        return self.get_cell_value(ExcelColumn.CASE_IF_EXECUTE, row)

    # 获取响应数据类型
    def get_data_type(self, row):
        return self.get_cell_value(ExcelColumn.CASE_DATA_TYPE, row)


if __name__ == '__main__':
    read = ReadExcel()
    # 实例化一个ReadExcel()类对象
    for row in range(2, read.get_row_count() + 1):
        # 从第二行开始遍历单元格，直至全部遍历完毕
        session = requests.session()
        # 实例化一个session对象，自动处理接口的cookie关联
        method = read.get_case_method(row)
        # 获取请求方法
        url = read.get_case_url(row)
        # 获取请求地址
        params = read.get_case_params_value(row)
        # 获取请求参数
        response = None
        # 定义接口响应变量，初始化为空
        if method == "GET":
            response = session.request(method=method, url=url, params=params)
        elif method == "POST":
            response = session.request(method=method, url=url, data=params)
        print(response)
        # 将响应数据以json格式编码返回输出，返回的数据类型为字典

        # http请求返回结果中文乱码问题
        # 1.问题描述：http请求中，请求的结果集中包含中文，最终以乱码展示
        # 2.问题本质：乱码的本质是服务端返回的字符集编码与客户端的编码方式不一致
        # 3.解决办法：让服务端返回的结果的编码与客服端的编码保持一致，最有效的方法是在request的Header中增加一项
        # Accept:application/json:charset=UTF-8
